import numpy as np
import random
# Generar archivo Excel con la solución
import openpyxl
from openpyxl.styles import Alignment
from collections import defaultdict



# Leer archivo de configuracion
def leer_configuracion(archivo):
    salones = {}
    profesores = {}
    materias = []

    with open(archivo, 'r') as file:
        seccion_actual = None

        for linea in file:
            linea = linea.strip()
            if not linea or linea.startswith('#'):
                continue

            if linea == '[Salones]':
                seccion_actual = 'Salones'
            elif linea == '[Profesores]':
                seccion_actual = 'Profesores'
            elif linea == '[Materias]':
                seccion_actual = 'Materias'
            else:
                if seccion_actual == 'Salones':
                    salon, horarios = linea.split(':')
                    salones[salon.strip()] = np.array(list(map(int, horarios.split(','))))

                elif seccion_actual == 'Profesores':
                    profesor, data = linea.split(':')
                    materias_horarios = data.split(';')
                    materias_profesor = materias_horarios[0].split(',')
                    horarios_profesor = np.array(list(map(int, materias_horarios[1].split(','))))
                    profesores[profesor.strip()] = {
                        'materias': materias_profesor,
                        'horarios': horarios_profesor
                    }

                elif seccion_actual == 'Materias':
                    materias.append(linea.strip())

    return salones, profesores, materias

# Generar poblacion inicial
def generar_poblacion_inicial(materias, salones, profesores, tamano_poblacion):
    poblacion = []

    for _ in range(tamano_poblacion):
        cromosoma = []
        ocupados = {"profesores": {}, "salones": {}}

        for materia in materias:
            profesores_validos = [p for p, datos in profesores.items() if materia in datos['materias']]
            if not profesores_validos:
                raise ValueError(f"No hay profesores disponibles para la materia {materia}")

            profesor = random.choice(profesores_validos)
            horarios_profesor = profesores[profesor]['horarios']

            horario_valido = False
            while not horario_valido:
                horario = random.choice(horarios_profesor)
                salones_validos = [s for s, h in salones.items() if horario in h]
                if not salones_validos:
                    continue
                salon = random.choice(salones_validos)

                if (profesor, horario) not in ocupados["profesores"] and (salon, horario) not in ocupados["salones"]:
                    cromosoma.append((materia, salon, profesor, horario))
                    ocupados["profesores"][(profesor, horario)] = True
                    ocupados["salones"][(salon, horario)] = True
                    horario_valido = True

        poblacion.append(cromosoma)

    return poblacion

# Funcion de aptitud
def calcular_aptitud(cromosoma):
    penalizacion = 0
    horarios_materia = {}
    horarios_profesor = {}
    horarios_salon = {}

    for materia, salon, profesor, horario in cromosoma:
        if(materia, horario) in horarios_materia:
            penalizacion += 1
        else:
            horarios_materia[(materia, horario)] = True

        if (profesor, horario) in horarios_profesor:
            penalizacion += 1
        else:
            horarios_profesor[(profesor, horario)] = True

        if (salon, horario) in horarios_salon:
            penalizacion += 1
        else:
            horarios_salon[(salon, horario)] = True

    return -penalizacion

# Seleccion por torneo
def seleccion_por_torneo(poblacion, aptitudes, k=2):
    seleccionados = np.random.choice(len(poblacion), k, replace=False)
    mejor = max(seleccionados, key=lambda i: aptitudes[i])
    return poblacion[mejor]

# Cruce de dos puntos
def cruce(padre1, padre2):
    
    punto1 = random.randint(0, len(padre1) - 1)
    punto2 = random.randint(punto1, len(padre1) - 1)

    hijo1 = padre1[:punto1] + padre2[punto1:punto2] + padre1[punto2:]
    hijo2 = padre2[:punto1] + padre1[punto1:punto2] + padre2[punto2:]

    return hijo1, hijo2

# Mutacion
def mutacion(cromosoma, salones, profesores):
    if random.random() < 0.01:
        indice = random.randint(0, len(cromosoma) - 1)
        materia, _, _, _ = cromosoma[indice]

        profesores_validos = [p for p, datos in profesores.items() if materia in datos['materias']]
        profesor = random.choice(profesores_validos)
        horario = random.choice(profesores[profesor]['horarios'])
        salones_validos = [s for s, h in salones.items() if horario in h]
        salon = random.choice(salones_validos)

        cromosoma[indice] = (materia, salon, profesor, horario)

# Algoritmo genetico
def algoritmo_genetico(salones, profesores, materias, tamano_poblacion=200, generaciones=1000):
    poblacion = generar_poblacion_inicial(materias, salones, profesores, tamano_poblacion)
    #print("\n")
    #print(poblacion[0])
    #print("\n")
    for generacion in range(generaciones):
        aptitudes = np.array([calcular_aptitud(cromosoma) for cromosoma in poblacion])
        nueva_poblacion = []

        while len(nueva_poblacion) < tamano_poblacion:
            padre1 = seleccion_por_torneo(poblacion, aptitudes)
            padre2 = seleccion_por_torneo(poblacion, aptitudes)

            hijo1, hijo2 = cruce(padre1, padre2)

            mutacion(hijo1, salones, profesores)
            mutacion(hijo2, salones, profesores)

            nueva_poblacion.extend([hijo1, hijo2])

        poblacion = nueva_poblacion[:tamano_poblacion]

        mejor_aptitud = aptitudes.max()
        print(f"Generacion {generacion + 1}: Mejor aptitud = {mejor_aptitud}")

    mejor_cromosoma = max(poblacion, key=calcular_aptitud)
    return mejor_cromosoma

def imprimir_solucion_tabla(mejor_solucion, salones, horarios_posibles, profesores):
    # Crear una estructura vacía para la tabla
    tabla = defaultdict(lambda: defaultdict(str))

    # Asignar las materias y profesores a los horarios y salones correspondientes
    for materia, salon, profesor, horario in mejor_solucion:
        profesor_nombre = profesor  # Obtener el nombre del profesor
        tabla[horario][salon] = f"{materia} ({profesor_nombre})"  # Incluir el nombre del profesor

    # Crear el archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Horarios Asignados"

    # Escribir los encabezados (salones)
    ws.cell(row=1, column=1, value="Hora/Sala")
    for col, salon in enumerate(salones.keys(), start=2):
        ws.cell(row=1, column=col, value=salon)

    # Escribir los horarios y las materias con el nombre del profesor
    for row, hora in enumerate(horarios_posibles, start=2):
        ws.cell(row=row, column=1, value=f"Hora {hora}")
        for col, salon in enumerate(salones.keys(), start=2):
            materia_con_profesor = tabla[hora].get(salon, '')
            ws.cell(row=row, column=col, value=materia_con_profesor)

    # Ajustar la alineación
    for row in ws.iter_rows(min_row=1, max_row=len(horarios_posibles) + 1, min_col=1, max_col=len(salones) + 1):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Guardar el archivo Excel
    wb.save("horarios_asignados_con_profesor.xlsx")
    print("Archivo Excel con la solución ha sido guardado como 'horarios_asignados.xlsx'.")



# Main
if __name__ == "__main__":
    archivo = 'configuracion.txt'
    salones, profesores, materias = leer_configuracion(archivo)
    print("Salones:", salones)
    print("Profesores:", profesores)
    print("Materias:", materias)


    horarios_posibles = sorted(set(h for horarios in salones.values() for h in horarios))

    mejor_solucion = algoritmo_genetico(salones, profesores, materias)
    
    print("\nMejor solución encontrada:")
    for asignacion in mejor_solucion:
        materia, salon, profesor, horario = asignacion
        print(f"[{materia}, {salon}, {profesor}, {int(horario)}]")

    imprimir_solucion_tabla(mejor_solucion, salones, horarios_posibles,profesores)
